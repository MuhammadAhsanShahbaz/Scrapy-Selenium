import csv
import json
import os

from openpyxl import load_workbook, Workbook


class Merging:
    def __init__(self, output_file):
        self.output_file = output_file
        self.average_price = None
        self.low_range = None
        self.high_range = None
        self.csv1_data = None
        self.csv1_headers = None
        self.excel1_data = None
        self.excel1_header = None
        self.length_before = None

    def load_json(self, json_file):
        with open(json_file, 'r') as file:
            json_data = json.load(file)
            try:
                self.average_price = int(float(json_data.get('Average Price Per Acre').replace('$', '').replace(',', '')))
                self.low_range = int(json_data.get('Low Range %').replace('%', ''))
                self.high_range = int(json_data.get('High Range %').replace('%', ''))
            except ValueError:
                self.average_price = 0
                self.low_range = 0
                self.high_range = 0

    def load_csv(self, csv_file):
        with open(csv_file, 'r', newline='') as file:
            csv_reader = csv.DictReader(file)
            self.csv1_headers = csv_reader.fieldnames[13:]  # Removing first 13 headers
            self.csv1_data = list(csv_reader)

    def load_excel(self, excel_file):
        work_book = load_workbook(excel_file)
        sheet = work_book.active

        self.excel1_header = [cell.value for cell in sheet[1]]
        self.excel1_header.insert(8, 'Acre')
        self.excel1_header.insert(9, 'lowrange')
        self.excel1_header.insert(10, 'highrange')
        self.length_before = len(self.excel1_header)

        self.excel1_data = {
            (row[0], row[9], row[10]): row for row in sheet.iter_rows(min_row=2, values_only=True)
        }

    def create_new_csv(self):
        with open(self.output_file, 'w', newline='') as csvfile:
            csv_writer = csv.writer(csvfile)
            headers = self.excel1_header + self.csv1_headers
            csv_writer.writerow(headers)

            self.write_data(csv_writer)

        with open(self.output_file, 'r', newline='') as csvfile:
            reader = csv.reader(csvfile)
            data = list(reader)

        columns_to_delete = []

        for col in range(len(data[0])):
            column_values = [row[col] for row in data[1:]]

            if all(value == '' for value in column_values):
                columns_to_delete.append(col)

        for col_index in reversed(columns_to_delete):
            for row in data:
                del row[col_index]

        with open(self.output_file, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerows(data)

    def write_data(self, csv_writer):
        for row in self.csv1_data:
            address = row.get('Input Property Address', '')
            first_name = row.get('Input First Name', '')
            last_name = row.get('Input Last Name', '')

            if (address, first_name, last_name) in self.excel1_data:
                excel_row = list(self.excel1_data[(address, first_name, last_name)])
                try:
                    data = self.get_acre_data(int(excel_row[7].replace(',', '').strip(' ')))
                except (TypeError, ValueError, AttributeError):
                    data = [0, 0, 0]

                for i, d in enumerate(data, start=8):
                    excel_row.insert(i, d)

                combined_row = list(excel_row) + [value for key, value in list(row.items())[13:]]
                csv_writer.writerow(combined_row)

    def get_acre_data(self, lot):
        range_data = [round(lot / 43560, 2)]
        ratio = (lot / 43560) * self.average_price
        range_data.append(f'${round(ratio * (self.low_range / 100), 2)}')
        range_data.append(f'${round(ratio * (self.high_range / 100), 2)}')
        return range_data

    def get_file_from_info(self):
        directory = 'Property info'
        files = os.listdir(directory)
        file_path = None

        for file in files:
            if file.endswith('.csv') or file.endswith('.xlsx'):
                file_path = os.path.join(directory, file)
                if file.endswith('.csv'):
                    excel_file_path = file_path.replace('.csv', '.xlsx')

                    wb = Workbook()
                    ws = wb.active

                    with open(file_path, newline='', encoding='utf-8') as csvfile:
                        csvreader = csv.reader(csvfile)

                        for row in csvreader:
                            ws.append(row)

                    wb.save(excel_file_path)

                    self.load_excel(excel_file_path)

                else:
                    self.load_excel(file_path)

                break

        if file_path is None:
            raise FileNotFoundError("No .csv or .xlsx file found in the input directory.")

    def get_file_from_owner_contact(self):
        directory = 'Property owner contact'
        files = os.listdir(directory)
        file_path = None

        for file in files:
            if file.endswith('.csv') or file.endswith('.xlsx'):
                file_path = os.path.join(directory, file)

                if file.endswith('.csv'):
                    self.load_csv(file_path)

                else:
                    csv_file_path = file_path.replace('.xlsx', '.csv')

                    wb = load_workbook(file_path)
                    ws = wb.active

                    with open(csv_file_path, 'w', newline='', encoding='utf-8') as csvfile:
                        csvwriter = csv.writer(csvfile)

                        for row in ws.iter_rows(values_only=True):
                            csvwriter.writerow(row)

                    self.load_csv(csv_file_path)

                break

        if file_path is None:
            raise FileNotFoundError("No .csv or .xlsx file found in the input directory.")

    def get_input_json_file(self):
        input_directory = os.getcwd()
        files = os.listdir(input_directory)
        file_path = None

        for file in files:
            if file.endswith('.json'):
                file_path = os.path.join(input_directory, file)
                self.load_json(file_path)
                break

        if file_path is None:
            raise FileNotFoundError("No .json file found in the current working directory.")


def main():
    merging = Merging('output.csv')
    merging.get_input_json_file()
    merging.get_file_from_info()
    merging.get_file_from_owner_contact()
    merging.create_new_csv()


if __name__ == "__main__":
    main()
