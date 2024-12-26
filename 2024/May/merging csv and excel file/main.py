import csv
import json
import os

from openpyxl import load_workbook
from copy import deepcopy


class Merger:
    def __init__(self):
        self.output_file = 'test9.csv'

        self.property_files = self.get_input_filenames('Property info')
        self.owner_files = self.get_input_filenames('Contact Info')

        self.property_info = self.get_data_from_file(self.property_files)
        self.owners_info = self.get_data_from_file(self.owner_files)

        self.user_inputs = self.get_input_json_file()
        self.output_headers = self.get_output_column_names()

        self.merge()

    def get_data_from_file(self, files_list):
        data = []

        for file in files_list:
            if file.endswith('.csv'):
                csv_data = self.read_csv(file)
                data += csv_data

            elif file.endswith('.xlsx'):
                excel_data = self.read_excel(file)
                data += excel_data

        return data

    def merge(self):
        print('\n\nMerging Data across files')
        owner_data = dict()
        merged_data = []

        for owner_row in self.owners_info:
            info_key = (owner_row.get('Input Property Address').strip() + owner_row
                        .get('Input First Name').strip() + owner_row.get('Input Last Name').strip())

            owner_data[info_key] = owner_row

        for prop_row in self.property_info:
            owner_key = (prop_row.get('Address').strip() + prop_row.get('Owner 1 First Name').strip()
                         + prop_row.get('Owner 1 Last Name').strip())
            owner_info = owner_data.get(owner_key)

            if not owner_info:
                continue

            prop_row.update(self.calculate_acre_prices(prop_row))
            prop_row.update(owner_info)

            merged_data.append(prop_row)

        print(f'{len(merged_data)} Rows Merged')

        unique_rows = self.remove_duplicates(merged_data)
        self.write_to_csv(unique_rows)

    def calculate_acre_prices(self, row):
        try:
            lot = int(row.get('Lot Size Sqft', '').replace(',', '').replace(' ', '').strip())

        except (ValueError, AttributeError):
            return {}

        per_acre_price = self.user_inputs.get('Average Price Per Acre')
        low_percentage = self.user_inputs.get('Low Range %')
        high_percentage = self.user_inputs.get('High Range %')

        acre_value = round(lot / 43560, 2)
        ratio = (lot / 43560) * per_acre_price
        low_range = f'${round(ratio * (low_percentage / 100), 2)}'
        high_range = f'${round(ratio * (high_percentage / 100), 2)}'

        market_value = f'${round(per_acre_price * acre_value, 2)}'

        return {'Acre': acre_value, 'Market Value': market_value, 'Low Range': low_range, 'High Range': high_range}

    def remove_empty_columns(self, rows):
        print(f'\n\nRemoving Empty Columns from the list')

        headers = deepcopy(self.output_headers)
        deleted_columns = []

        for col in headers:
            column_values = [row.get(col) for row in rows]

            if all(not value for value in column_values):
                deleted_columns.append(col)
                self.output_headers.remove(col)

        print(f'{len(deleted_columns)} Columns Removed as they were Empty')

        # print('\nRemoved Columns:')
        # for de in deleted_columns:
        #     print(de)

    def remove_duplicates(self, rows):
        unique_rows = []
        seen_rows = set()

        print('\n\bRemoving duplicates:')
        duplciates_removed_count = 0

        for row in rows:
            row_data = ''.join([str(v) for v in row.values()])

            if row_data in seen_rows:
                duplciates_removed_count += 1
                continue

            seen_rows.add(row_data)
            unique_rows.append(row)

        print(f'{duplciates_removed_count} Duplicate Rows Removed')
        return unique_rows

    def write_to_csv(self, data):
        self.remove_empty_columns(data)

        print('\n\nWriting Results To CSV file')

        with open(self.output_file, 'w', newline='') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=self.output_headers)
            writer.writeheader()

            for row in data:
                filtered_row = {header: row.get(header, '') for header in self.output_headers}
                writer.writerow(filtered_row)

        print(f'{len(data)} Rows written to the CSV file: "{self.output_file}"\n\n')

    def get_input_filenames(self, dir_path):
        file_names = []

        files = os.listdir(dir_path)

        for file in files:
            if file.endswith('.csv') or file.endswith('.xlsx'):
                file_path = os.path.join(dir_path, file)
                file_names.append(file_path)

        print(f'\n\n{len(file_names)} CSV/Excel files found in the directory: "{dir_path}"')

        return file_names

    def read_csv(self, filename):
        data = []

        print(f'\n\nGetting data from CSV file: "{filename}"')

        with open(filename, 'r', newline='') as csvfile:
            reader = csv.DictReader(csvfile)

            for row in reader:
                data.append(row)

        print(f'{len(data)} Rows fetched')
        return data

    def read_excel(self, filename):
        data = []
        print(f'\n\nGetting data from Excel file: "{filename}"')

        workbook = load_workbook(filename)
        sheet = workbook.active

        headers = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            data.append(row_dict)

        print(f'{len(data)} Rows fetched')
        return data

    def get_input_json_file(self):
        with open('input.json', 'r') as json_file:
            json_data = json.load(json_file)
            input_data = {}

            try:
                input_data['Average Price Per Acre'] = int(
                    float(json_data.get('Average Price Per Acre').replace('$', '').replace(',', '')))
                input_data['Low Range %'] = int(json_data.get('Low Range %').replace('%', ''))
                input_data['High Range %'] = int(json_data.get('High Range %').replace('%', ''))
                return input_data

            except ValueError:
                return input_data

    def get_output_column_names(self):
        property_headers = []

        for data in self.property_info:
            for d in data.keys():
                if d not in property_headers:
                    property_headers.append(d)

        list_to_insert = ["Acre", "Market Value", "Low Range", "High Range"]
        property_headers = property_headers[:8] + list_to_insert + property_headers[8:]

        contract_headers = []

        for data in self.owners_info:
            for d in data.keys():
                if d not in contract_headers:
                    if 'Input' not in d:
                        contract_headers.append(d)

        return property_headers + contract_headers


if __name__ == "__main__":
    print('\n\nStarted...')
    merger = Merger()

    print('\n\nFinished.')

    os.system("pause")  # pause until user press any key to terminate the window
