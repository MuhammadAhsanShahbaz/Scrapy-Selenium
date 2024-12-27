import csv
import json
import os
from copy import deepcopy

from openpyxl import load_workbook


class Merger:
    def __init__(self, output_file):
        self.output_file = output_file

        self.user_inputs = self.get_input_json_file()
        self.output_headers = ["Address", "Unit #", "City", "State", "Zip", "County", "APN", "Lot Size Sqft", "Acre",
                               "Low Range", "High Range", "Owner Occupied", "Owner 1 First Name", "Owner 1 Last Name",
                               "Owner 2 First Name", "Owner 2 Last Name", "Mailing Care of Name", "Mailing Address",
                               "Mailing Unit #", "Mailing City", "Mailing State", "Mailing Zip", "Mailing County",
                               "Do Not Mail", "Property Type", "Bedrooms", "Marketing Lists", "Total Bathrooms",
                               "Building Sqft", "Effective Year Built", "Total Assessed Value",
                               "Last Sale Recording Date", "Last Sale Amount", "Total Open Loans",
                               "Est. Remaining balance of Open Loans", "Est. Value", "Est. Loan-to-Value",
                               "Est. Equity", "MLS Status", "MLS Date", "MLS Amount", "Lien Amount",
                               "Date Added to List", "Method of Add", "Input Last Name", "Input First Name",
                               "Input Mailing Address", "Input Mailing City", "Input Mailing State",
                               "Input Mailing Zip", "Input Property Address", "Input Property City",
                               "Input Property State", "Input Property Zip", "Input Custom Field 1",
                               "Input Custom Field 2", "Input Custom Field 3", "Owner Fix - Last Name",
                               "Owner Fix - First Name", "Owner Fix - Mailing Address", "Owner Fix - Mailing City",
                               "Owner Fix - Mailing State", "Owner Fix - Mailing Zip", "ResultCode",
                               "Matched First Name", "Matched Last Name", "DNC/Litigator Scrub", "Age", "Deceased",
                               "Phone1", "Phone1 Type", "Phone2", "Phone2 Type", "Phone3", "Phone3 Type", "Phone4",
                               "Phone4 Type", "Phone5", "Phone5 Type", "Phone6", "Phone6 Type", "Phone7", "Phone7 Type",
                               "Email1", "Email2", "Confirmed Mailing Address", "Confirmed Mailing City",
                               "Confirmed Mailing State", "Confirmed Mailing Zip", "Relative1 Name", "Relative1 Age",
                               "Relative1 Phone1", "Relative1 Phone1 Type", "Relative1 Phone2", "Relative1 Phone2 Type",
                               "Relative1 Phone3", "Relative1 Phone3 Type", "Relative1 Phone4", "Relative1 Phone4 Type",
                               "Relative1 Phone5", "Relative1 Phone5 Type", "Relative2 Name", "Relative2 Age",
                               "Relative2 Phone1", "Relative2 Phone1 Type", "Relative2 Phone2", "Relative2 Phone2 Type",
                               "Relative2 Phone3", "Relative2 Phone3 Type", "Relative2 Phone4", "Relative2 Phone4 Type",
                               "Relative2 Phone5", "Relative2 Phone5 Type", "Relative3 Name", "Relative3 Age",
                               "Relative3 Phone1", "Relative3 Phone1 Type", "Relative3 Phone2", "Relative3 Phone2 Type",
                               "Relative3 Phone3", "Relative3 Phone3 Type", "Relative3 Phone4", "Relative3 Phone4 Type",
                               "Relative3 Phone5", "Relative3 Phone5 Type", "Relative4 Name", "Relative4 Age",
                               "Relative4 Phone1", "Relative4 Phone1 Type", "Relative4 Phone2", "Relative4 Phone2 Type",
                               "Relative4 Phone3", "Relative4 Phone3 Type", "Relative4 Phone4", "Relative4 Phone4 Type",
                               "Relative4 Phone5", "Relative4 Phone5 Type", "Relative5 Name", "Relative5 Age",
                               "Relative5 Phone1", "Relative5 Phone1 Type", "Relative5 Phone2", "Relative5 Phone2 Type",
                               "Relative5 Phone3", "Relative5 Phone3 Type", "Relative5 Phone4", "Relative5 Phone4 Type",
                               "Relative5 Phone5", "Relative5 Phone5 Type", "Person2 First Name", "Person2 Last Name",
                               "Person2 Age", "Person2 Deceased", "Person2 Phone1", "Person2 Phone1 Type",
                               "Person2 Phone2", "Person2 Phone2 Type", "Person2 Phone3", "Person2 Phone3 Type",
                               "Person2 Phone4", "Person2 Phone4 Type", "Person2 Phone5", "Person2 Phone5 Type",
                               "Person2 Phone6", "Person2 Phone6 Type", "Person2 Phone7", "Person2 Phone7 Type",
                               "Person2 Email1", "Person2 Email2", "Person2 Confirmed Mailing Address",
                               "Person2 Confirmed Mailing City", "Person2 Confirmed Mailing State",
                               "Person2 Confirmed Mailing Zip", "Person2 Relative1 Name", "Person2 Relative1 Age",
                               "Person2 Relative1 Phone1", "Person2 Relative1 Phone1 Type", "Person2 Relative1 Phone2",
                               "Person2 Relative1 Phone2 Type", "Person2 Relative1 Phone3",
                               "Person2 Relative1 Phone3 Type", "Person2 Relative1 Phone4",
                               "Person2 Relative1 Phone4 Type", "Person2 Relative1 Phone5",
                               "Person2 Relative1 Phone5 Type", "Person2 Relative2 Name",
                               "Person2 Relative2 Age",
                               "Person2 Relative2 Phone1", "Person2 Relative2 Phone1 Type", "Person2 Relative2 Phone2",
                               "Person2 Relative2 Phone2 Type",
                               "Person2 Relative2 Phone3", "Person2 Relative2 Phone3 Type", "Person2 Relative2 Phone4",
                               "Person2 Relative2 Phone4 Type",
                               "Person2 Relative2 Phone5", "Person2 Relative2 Phone5 Type", "Person2 Relative3 Name",
                               "Person2 Relative3 Age",
                               "Person2 Relative3 Phone1", "Person2 Relative3 Phone1 Type", "Person2 Relative3 Phone2",
                               "Person2 Relative3 Phone2 Type",
                               "Person2 Relative3 Phone3", "Person2 Relative3 Phone3 Type", "Person2 Relative3 Phone4",
                               "Person2 Relative3 Phone4 Type",
                               "Person2 Relative3 Phone5", "Person2 Relative3 Phone5 Type", "Person2 Relative4 Name",
                               "Person2 Relative4 Age",
                               "Person2 Relative4 Phone1", "Person2 Relative4 Phone1 Type", "Person2 Relative4 Phone2",
                               "Person2 Relative4 Phone2 Type",
                               "Person2 Relative4 Phone3", "Person2 Relative4 Phone3 Type", "Person2 Relative4 Phone4",
                               "Person2 Relative4 Phone4 Type",
                               "Person2 Relative4 Phone5", "Person2 Relative4 Phone5 Type", "Person2 Relative5 Name",
                               "Person2 Relative5 Age",
                               "Person2 Relative5 Phone1", "Person2 Relative5 Phone1 Type", "Person2 Relative5 Phone2",
                               "Person2 Relative5 Phone2 Type",
                               "Person2 Relative5 Phone3", "Person2 Relative5 Phone3 Type", "Person2 Relative5 Phone4",
                               "Person2 Relative5 Phone4 Type",
                               "Person2 Relative5 Phone5", "Person2 Relative5 Phone5 Type", "Person3 First Name",
                               "Person3 Last Name",
                               "Person3 Age", "Person3 Deceased", "Person3 Phone1", "Person3 Phone1 Type",
                               "Person3 Phone2", "Person3 Phone2 Type",
                               "Person3 Phone3", "Person3 Phone3 Type", "Person3 Phone4", "Person3 Phone4 Type",
                               "Person3 Phone5", "Person3 Phone5 Type",
                               "Person3 Phone6", "Person3 Phone6 Type", "Person3 Phone7", "Person3 Phone7 Type",
                               "Person3 Email1", "Person3 Email2",
                               "Person3 Confirmed Mailing Address", "Person3 Confirmed Mailing City",
                               "Person3 Confirmed Mailing State",
                               "Person3 Confirmed Mailing Zip", "Person3 Relative1 Name", "Person3 Relative1 Age",
                               "Person3 Relative1 Phone1",
                               "Person3 Relative1 Phone1 Type", "Person3 Relative1 Phone2",
                               "Person3 Relative1 Phone2 Type", "Person3 Relative1 Phone3",
                               "Person3 Relative1 Phone3 Type", "Person3 Relative1 Phone4",
                               "Person3 Relative1 Phone4 Type", "Person3 Relative1 Phone5",
                               "Person3 Relative1 Phone5 Type", "Person3 Relative2 Name", "Person3 Relative2 Age",
                               "Person3 Relative2 Phone1",
                               "Person3 Relative2 Phone1 Type", "Person3 Relative2 Phone2",
                               "Person3 Relative2 Phone2 Type", "Person3 Relative2 Phone3",
                               "Person3 Relative2 Phone3 Type", "Person3 Relative2 Phone4",
                               "Person3 Relative2 Phone4 Type", "Person3 Relative2 Phone5",
                               "Person3 Relative2 Phone5 Type", "Person3 Relative3 Name", "Person3 Relative3 Age",
                               "Person3 Relative3 Phone1",
                               "Person3 Relative3 Phone1 Type", "Person3 Relative3 Phone2",
                               "Person3 Relative3 Phone2 Type", "Person3 Relative3 Phone3",
                               "Person3 Relative3 Phone3 Type", "Person3 Relative3 Phone4",
                               "Person3 Relative3 Phone4 Type", "Person3 Relative3 Phone5",
                               "Person3 Relative3 Phone5 Type", "Person3 Relative4 Name", "Person3 Relative4 Age",
                               "Person3 Relative4 Phone1",
                               "Person3 Relative4 Phone1 Type", "Person3 Relative4 Phone2",
                               "Person3 Relative4 Phone2 Type", "Person3 Relative4 Phone3",
                               "Person3 Relative4 Phone3 Type", "Person3 Relative4 Phone4",
                               "Person3 Relative4 Phone4 Type", "Person3 Relative4 Phone5",
                               "Person3 Relative4 Phone5 Type", "Person3 Relative5 Name", "Person3 Relative5 Age",
                               "Person3 Relative5 Phone1",
                               "Person3 Relative5 Phone1 Type", "Person3 Relative5 Phone2",
                               "Person3 Relative5 Phone2 Type", "Person3 Relative5 Phone3",
                               "Person3 Relative5 Phone3 Type", "Person3 Relative5 Phone4",
                               "Person3 Relative5 Phone4 Type", "Person3 Relative5 Phone5",
                               "Person3 Relative5 Phone5 Type"]

        self.property_files = self.get_input_filenames('Property info')
        self.owner_files = self.get_input_filenames('Property owner contact')

        self.property_info = self.get_data_from_file(self.property_files)
        self.owners_info = self.get_data_from_file(self.owner_files)

        self.merge()

    def get_data_from_file(self, files_list):
        data = []

        for file in files_list:
            if file.endswith('.csv'):
                csv_data = self.read_csv(file)
                data += csv_data

            elif file.endswith('.xlsx'):
                excel_data = self.read_excel(file)
                data += excel_data

        return data

    def merge(self):
        owner_data = dict()
        merged_data = []

        for owner_row in self.owners_info:
            info_key = (owner_row.get('Input Property Address', '').strip() + owner_row
                        .get('Input First Name', '').strip() + owner_row.get('Input Last Name', '').strip())

            owner_data[info_key] = owner_row

        for prop_row in self.property_info:
            owner_key = (prop_row.get('Address', '').strip() + prop_row.get('Owner 1 First Name', '').strip()
                         + prop_row.get('Owner 1 Last Name', '').strip())
            owner_info = owner_data.get(owner_key)

            if not owner_info:
                continue

            prop_row.update(self.calculate_acre_price(prop_row))
            prop_row.update(owner_info)

            merged_data.append(prop_row)

        unique_rows = self.remove_duplicates(merged_data)
        self.write_to_csv(unique_rows)

    def calculate_acre_price(self, row):
        try:
            lot = int(row.get('Lot Size Sqft', '').replace(',', '').replace(' ', '').strip())

        except (ValueError, AttributeError):
            return {}

        per_acre_price = self.user_inputs.get('Average Price Per Acre', '')
        low_percentage = self.user_inputs.get('Low Range %', '')
        high_percentage = self.user_inputs.get('High Range %', '')

        acre_value = round(lot / 43560, 2)
        ratio = (lot / 43560) * per_acre_price
        low_range = f'${round(ratio * (low_percentage / 100), 2)}'
        high_range = f'${round(ratio * (high_percentage / 100), 2)}'

        return {'Acre': acre_value, 'Low Range': low_range, 'High Range': high_range}

    def remove_empty_columns(self, rows):
        headers = deepcopy(self.output_headers)
        deleted = []

        for col in headers:
            column_values = [row.get(col) for row in rows]

            if all(not value for value in column_values):
                deleted.append(col)
                self.output_headers.remove(col)

        if deleted:
            print('Columns are deleted: ')
            for de in deleted:
                print(de)

    def remove_duplicates(self, rows):
        unique_rows = []
        seen_rows = set()

        print('Removing duplicated data:')
        duplicates_removed_count = 0

        for row in rows:
            row_data = ''.join([str(v) for v in row.values()])

            if row_data in seen_rows:
                duplicates_removed_count += 1
                continue

            seen_rows.add(row_data)
            unique_rows.append(row)

        print(f'{duplicates_removed_count} are removed')
        return unique_rows

    def write_to_csv(self, data):
        self.remove_empty_columns(data)

        with open(self.output_file, 'w', newline='') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=self.output_headers)
            writer.writeheader()

            for row in data:
                filtered_row = {header: row.get(header, '') for header in self.output_headers}
                writer.writerow(filtered_row)

    def get_input_filenames(self, dir_path):
        file_names = []

        files = os.listdir(dir_path)

        for file in files:
            if file.endswith('.csv') or file.endswith('.xlsx'):
                file_path = os.path.join(dir_path, file)
                file_names.append(file_path)

        return file_names

    def read_csv(self, filename):
        data = []
        print(f'Getting data from {filename}')
        with open(filename, 'r', newline='') as csvfile:
            reader = csv.DictReader(csvfile)

            for row in reader:
                data.append(row)

        print(f'{len(data)} lines get fetched')
        return data

    def read_excel(self, filename):
        data = []
        print(f'Getting data from {filename}')

        workbook = load_workbook(filename)
        sheet = workbook.active

        headers = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            data.append(row_dict)

        print(f'{len(data)} lines get fetched')
        return data

    def get_input_json_file(self):
        with open('input.json', 'r') as json_file:
            json_data = json.load(json_file)
            input_data = {}

            try:
                input_data['Average Price Per Acre'] = int(
                            float(json_data.get('Average Price Per Acre').replace('$', '').replace(',', '')))
                input_data['Low Range %'] = int(json_data.get('Low Range %').replace('%', ''))
                input_data['High Range %'] = int(json_data.get('High Range %').replace('%', ''))
                return input_data

            except ValueError:
                return input_data


def main():
    merging = Merger('output.csv')


if __name__ == "__main__":
    main()
