import csv
import os

from openpyxl import load_workbook


class Merger:
    def __init__(self):
        self.output_file = 'output/test1.csv'

        self.input = self.get_input_filenames('input')
        self.input_info = self.get_data_from_file(self.input)

        self.output_headers = self.get_output_column_names()

        self.merge()

    def get_data_from_file(self, files_list):
        data = []

        for file in files_list:
            if file.endswith('.csv'):
                csv_data = self.read_csv(file)
                data += csv_data

            elif file.endswith('.xlsx'):
                excel_data = self.read_excel(file)
                data += excel_data

        return data

    def read_csv(self, filename):
        data = []

        print(f'\n\nGetting data from CSV file: "{filename}"')

        with open(filename, 'r', newline='') as csvfile:
            reader = csv.DictReader(csvfile)

            for row in reader:
                data.append(row)

        print(f'{len(data)} Rows fetched')

        return data

    def read_excel(self, filename):
        data = []
        print(f'\n\nGetting data from Excel file: "{filename}"')

        workbook = load_workbook(filename)
        sheet = workbook.active

        headers = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            data.append(row_dict)

        print(f'{len(data)} Rows fetched')
        return data

    def get_input_filenames(self, dir_path):
        file_names = []

        files = os.listdir(dir_path)

        for file in files:
            if file.endswith('.csv') or file.endswith('.xlsx'):
                file_path = os.path.join(dir_path, file)
                file_names.append(file_path)

        print(f'\n\n{len(file_names)} CSV/Excel files found in the directory: "{dir_path}"')

        return file_names

    def merge(self):
        apn_id = []
        final_list = []
        duplicate_count = 0

        for row in self.input_info:
            if row.get('APN') not in apn_id:
                apn_id.append(row.get('APN'))
                final_list.append(row)
            else:
                duplicate_count += 1

        print(f'\n\n{duplicate_count} duplication is/are found in the files')

        self.write_to_csv(final_list)

    def get_output_column_names(self):
        headers = []

        for data in self.input_info:
            for d in data.keys():
                if d not in headers:
                    headers.append(d)

        return headers

    def write_to_csv(self, data):
        with open(self.output_file, 'a', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=self.output_headers)

            if csvfile.tell() == 0:
                writer.writeheader()

            for row in data:
                writer.writerow(row)


def main():
    merger = Merger()


if __name__ == "__main__":
    main()
